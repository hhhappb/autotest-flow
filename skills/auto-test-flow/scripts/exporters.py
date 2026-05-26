#!/usr/bin/env python3
"""Export generated test cases to spreadsheet and mind-map formats."""

import json
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def write_test_cases_xlsx(path: Path, test_cases: dict) -> None:
    """Write test cases as an Excel workbook using openpyxl."""
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _test_case_export_rows(test_cases)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(rows[0])
    ws.row_dimensions[1].font = Font(bold=True)
    for row in rows[1:]:
        ws.append(row)

    column_widths = [16, 28, 14, 14, 14, 36, 36, 36, 36]
    for col_index, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    wb.save(path)


def write_test_cases_xmind(path: Path, test_cases: dict) -> None:
    """Write test cases as a simple XMind 2020 compatible workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)
    feature = str(test_cases.get("feature") or "自动化测试用例")
    attached = []
    for index, case in enumerate(test_cases.get("cases", []), start=1):
        title = f"{case.get('id', f'TC_{index:03d}')} {case.get('title', '')}".strip()
        children = []
        for label, key in [
            ("前置条件", "preconditions"),
            ("测试步骤", "steps"),
            ("测试数据", "test_data"),
            ("预期结果", "expected_result"),
        ]:
            children.append({
                "id": f"case-{index}-{key}",
                "class": "topic",
                "title": f"{label}: {_plain_text(case.get(key))}",
            })
        attached.append({
            "id": f"case-{index}",
            "class": "topic",
            "title": title,
            "labels": [str(case.get("priority") or ""), str(case.get("type") or "")],
            "children": {"attached": children},
        })
    content = [{
        "id": "sheet-1",
        "class": "sheet",
        "title": feature,
        "rootTopic": {
            "id": "root",
            "class": "topic",
            "title": feature,
            "children": {"attached": attached},
        },
    }]
    metadata = {
        "creator": {"name": "auto-test-flow"},
        "activeSheetId": "sheet-1",
    }
    manifest = {
        "file-entries": {
            "content.json": {},
            "metadata.json": {},
        }
    }
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("content.json", json.dumps(content, ensure_ascii=False, indent=2))
        archive.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False, indent=2))
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))


def _test_case_export_rows(test_cases: dict) -> list[list[str]]:
    rows = [[
        "用例编号",
        "用例标题",
        "优先级",
        "类型",
        "自动化候选",
        "前置条件",
        "测试步骤",
        "测试数据",
        "预期结果",
    ]]
    for case in test_cases.get("cases", []):
        rows.append([
            _plain_text(case.get("id")),
            _plain_text(case.get("title")),
            _plain_text(case.get("priority")),
            _plain_text(case.get("type")),
            _plain_text(case.get("automation_candidate")),
            _plain_text(case.get("preconditions")),
            _plain_text(case.get("steps")),
            _plain_text(case.get("test_data")),
            _plain_text(case.get("expected_result")),
        ])
    return rows


def _plain_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    if isinstance(value, dict):
        return "\n".join(f"{key}: {val}" for key, val in value.items())
    return str(value)
