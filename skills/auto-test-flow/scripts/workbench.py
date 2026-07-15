#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Local web workbench for auto-test-flow.

The workbench serves a local web UI, runs the existing orchestrator,
and can hand a generated codex_task.md to `codex.cmd exec` while
keeping logs beside the generated run artifacts.
"""

from __future__ import annotations

import argparse
import base64
import csv
import html
import io
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import uuid
import webbrowser
from datetime import datetime
from pathlib import Path, PurePosixPath
from urllib.parse import quote, unquote

import openpyxl
import uvicorn
from fastapi import FastAPI, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles

from workbench_cdp import (
    DEFAULT_ELECTRON_DEBUG_PORT,
    DEFAULT_STORE_DEBUG_PORT,
    prepare_authorized_cdp_environment,
    resolve_test_project_root,
)
from workbench_codex import start_codex_execution_job
from workbench_evidence import (
    EVIDENCE_MODES,
    maybe_write_failure_evidence_diff,
    start_evidence_job as start_evidence_capture_job,
)


FINAL_STATUSES = {"success", "failed"}
REVIEW_POLICIES = {"auto-review", "full-auto"}
APPROVAL_POLICIES = {"on-request", "never"}
EXECUTION_MODES = {"analysis", "authorized"}
DEEPSEEK_MODELS = {"deepseek-v4-flash", "deepseek-v4-pro"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
TEXT_EXTENSIONS = {".txt", ".md", ".csv"}
SPREADSHEET_EXTENSIONS = {".xlsx", ".xls"}
SUPPORTED_ATTACHMENT_EXTENSIONS = IMAGE_EXTENSIONS | TEXT_EXTENSIONS | SPREADSHEET_EXTENSIONS
MAX_JOB_LOG_LINES = 3000


class WorkbenchState:
    """Shared state for the local HTTP handler."""

    def __init__(
        self,
        host: str,
        port: int,
        orchestrator_path: Path,
        output_dir: Path,
        project_root: Path,
        python_executable: str,
        test_python_executable: str,
        explicit_test_python: str,
        codex_command: str,
        model: str = "deepseek-v4-flash",
    ) -> None:
        self.host = host
        self.port = port
        self.orchestrator_path = orchestrator_path.resolve()
        self.output_dir = output_dir.resolve()
        self.project_root = project_root.resolve()
        self.python_executable = python_executable
        self.test_python_executable = test_python_executable
        self.explicit_test_python = explicit_test_python
        self.codex_command = codex_command
        self.model = model
        self.jobs: dict[str, dict] = {}
        self.processes: dict[str, subprocess.Popen] = {}
        self.lock = threading.Lock()
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def create_job(self, kind: str) -> dict:
        job_id = uuid.uuid4().hex[:12]
        job = {
            "id": job_id,
            "kind": kind,
            "status": "running",
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "command": [],
            "logs": [],
            "run_dir": None,
            "run_url": None,
            "log_file": None,
            "last_message_file": None,
            "summary_file": None,
            "exit_code": None,
            "error": None,
            "interactive": False,
        }
        with self.lock:
            self.jobs[job_id] = job
        return job

    def append_log(self, job_id: str, line: str) -> None:
        with self.lock:
            job = self.jobs[job_id]
            job["logs"].append(line)
            job["logs"] = job["logs"][-MAX_JOB_LOG_LINES:]
            job["updated_at"] = datetime.now().isoformat(timespec="seconds")

    def update_job(self, job_id: str, **fields) -> None:
        with self.lock:
            job = self.jobs[job_id]
            job.update(fields)
            job["updated_at"] = datetime.now().isoformat(timespec="seconds")

    def get_job(self, job_id: str) -> dict | None:
        with self.lock:
            job = self.jobs.get(job_id)
            snapshot = json.loads(json.dumps(job, ensure_ascii=False)) if job else None
        if not snapshot:
            return None

        file_logs = self._read_log_file_tail(snapshot.get("log_file"))
        if file_logs:
            if snapshot.get("status") != "running":
                memory_logs = snapshot.get("logs") or []
                for line in memory_logs:
                    if line not in file_logs:
                        file_logs.append(line)
            snapshot["logs"] = file_logs[-MAX_JOB_LOG_LINES:]
        return snapshot

    @staticmethod
    def _read_log_file_tail(log_file: str | None) -> list[str]:
        if not log_file:
            return []
        path = Path(log_file)
        if not path.exists() or not path.is_file():
            return []
        try:
            return path.read_text(encoding="utf-8", errors="replace").splitlines()[-MAX_JOB_LOG_LINES:]
        except OSError:
            return []

    def attach_process(self, job_id: str, process: subprocess.Popen) -> None:
        with self.lock:
            self.processes[job_id] = process

    def detach_process(self, job_id: str) -> None:
        with self.lock:
            self.processes.pop(job_id, None)
            if job_id in self.jobs:
                self.jobs[job_id]["interactive"] = False

    def send_input(self, job_id: str, text: str) -> dict:
        with self.lock:
            job = self.jobs.get(job_id)
            process = self.processes.get(job_id)
            if not job or not process or process.stdin is None or process.poll() is not None:
                raise ValueError("No running interactive process for this job")
            process.stdin.write(text.rstrip("\n") + "\n")
            process.stdin.flush()
            job["logs"].append(f">>> {text.rstrip()}")
            job["updated_at"] = datetime.now().isoformat(timespec="seconds")
            return json.loads(json.dumps(job, ensure_ascii=False))

    def list_runs(self) -> list[dict]:
        runs = []
        for path in sorted(self.output_dir.glob("*"), key=lambda p: p.stat().st_mtime, reverse=True):
            if not path.is_dir() or not (path / "index.html").exists():
                continue
            review = self._read_run_review(path)
            codex = self._read_codex_status(path)
            runs.append(
                {
                    "name": path.name,
                    "path": str(path),
                    "url": f"/runs/{quote(path.name)}/index.html",
                    "review_url": f"/runs/{quote(path.name)}/index.html#doc-md-review-notes-md",
                    "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                    "review_decision": review.get("decision"),
                    "review_summary": review.get("summary"),
                    "review_counts": review.get("counts", {}),
                    "codex_status": codex.get("status"),
                    "codex_exit_code": codex.get("exit_code"),
                    "codex_updated_at": codex.get("updated_at"),
                    "codex_summary_url": codex.get("summary_url"),
                    "codex_log_url": codex.get("log_url"),
                    "codex_last_message_url": codex.get("last_message_url"),
                    "evidence_url": self._run_file_url(path, "md/evidence.md"),
                    "evidence_diff_url": self._run_file_url(path, "md/evidence_diff.md"),
                }
            )
        return runs[:50]

    def delete_run(self, run_ref: str) -> None:
        run_dir = self.resolve_run_dir(run_ref)
        output_root = self.output_dir.resolve()
        if run_dir == output_root or output_root not in run_dir.parents:
            raise ValueError("run_dir must be inside the configured output directory")
        shutil.rmtree(run_dir)

    @staticmethod
    def _read_run_review(run_dir: Path) -> dict:
        review_path = run_dir / "json" / "review_result.json"
        if review_path.exists():
            try:
                payload = json.loads(review_path.read_text(encoding="utf-8"))
                if isinstance(payload, dict):
                    return payload
            except (OSError, json.JSONDecodeError):
                return {}

        codex_task = run_dir / "md" / "codex_task.md"
        if codex_task.exists():
            try:
                text = codex_task.read_text(encoding="utf-8", errors="replace").lower()
            except OSError:
                text = ""
            if "skipped by the review gate" in text:
                return {
                    "decision": "blocked",
                    "summary": "Codex handoff was skipped by the review gate.",
                    "counts": {},
                }
        return {}

    @staticmethod
    def _run_file_url(run_dir: Path, relative_path: str) -> str | None:
        if not (run_dir / relative_path).exists():
            return None
        return f"/runs/{quote(run_dir.name)}/{relative_path}"

    @staticmethod
    def _read_codex_status(run_dir: Path) -> dict:
        logs_dir = run_dir / "logs"
        if not logs_dir.exists():
            return {"status": "not_started"}

        decision_files = sorted(logs_dir.glob("codex_decision_*.md"), key=lambda p: p.stat().st_mtime, reverse=True)
        exec_files = sorted(logs_dir.glob("codex_exec_*.log"), key=lambda p: p.stat().st_mtime, reverse=True)
        last_message_files = sorted(
            logs_dir.glob("codex_last_message_*.md"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if not decision_files and not exec_files and not last_message_files:
            return {"status": "not_started"}

        def codex_job_id(path: Path) -> str:
            stem = path.stem
            for prefix in ("codex_decision_", "codex_exec_", "codex_last_message_"):
                if stem.startswith(prefix):
                    return stem[len(prefix):]
            return ""

        latest_file = max(
            [*decision_files, *exec_files, *last_message_files],
            key=lambda p: p.stat().st_mtime,
        )
        current_job_id = codex_job_id(latest_file)
        if not current_job_id:
            return {"status": "unknown"}

        decision_file = logs_dir / f"codex_decision_{current_job_id}.md"
        exec_file = logs_dir / f"codex_exec_{current_job_id}.log"
        last_message_file = logs_dir / f"codex_last_message_{current_job_id}.md"
        decision_file = decision_file if decision_file.exists() else None
        exec_file = exec_file if exec_file.exists() else None
        last_message_file = last_message_file if last_message_file.exists() else None

        exit_code = None
        if decision_file:
            try:
                decision_text = decision_file.read_text(encoding="utf-8", errors="replace")
                match = re.search(r"(?:退出码|Exit code)\s*[:：]\s*(-?\d+)", decision_text)
                if match:
                    exit_code = int(match.group(1))
            except OSError:
                exit_code = None
        elif exec_file:
            try:
                tail_text = exec_file.read_text(encoding="utf-8", errors="replace")[-4000:]
                match = re.search(r"Exit code\s*[:：]\s*(-?\d+)", tail_text)
                if match:
                    exit_code = int(match.group(1))
            except OSError:
                exit_code = None

        if exit_code == 0:
            status = "success"
        elif exit_code is None:
            status = "unknown"
        else:
            status = "failed"

        latest_files = [path for path in (decision_file, exec_file, last_message_file) if path and path.exists()]
        latest_mtime = max((path.stat().st_mtime for path in latest_files), default=None)

        def log_url(path: Path | None) -> str | None:
            if not path:
                return None
            return f"/runs/{quote(run_dir.name)}/logs/{quote(path.name)}"

        return {
            "status": status,
            "exit_code": exit_code,
            "updated_at": datetime.fromtimestamp(latest_mtime).isoformat(timespec="seconds") if latest_mtime else None,
            "summary_url": log_url(decision_file),
            "log_url": log_url(exec_file),
            "last_message_url": log_url(last_message_file),
        }

    def resolve_run_dir(self, run_ref: str) -> Path:
        candidate = Path(run_ref)
        if not candidate.is_absolute():
            candidate = self.output_dir / run_ref
        candidate = candidate.resolve()
        output_root = self.output_dir.resolve()
        if candidate != output_root and output_root not in candidate.parents:
            raise ValueError("run_dir must be inside the configured output directory")
        if not candidate.exists() or not candidate.is_dir():
            raise FileNotFoundError(f"Run folder does not exist: {candidate}")
        if not (candidate / "index.html").exists():
            raise FileNotFoundError(f"Run folder has no index.html: {candidate}")
        return candidate

    def update_settings(self, project_root: Path, output_dir: Path, model: str | None = None) -> None:
        self.project_root = project_root.resolve()
        self.output_dir = output_dir.resolve()
        if model:
            self.model = model
        self.output_dir.mkdir(parents=True, exist_ok=True)


def run_command(
    state: WorkbenchState,
    job_id: str,
    command: list[str],
    cwd: Path,
    log_path: Path,
    input_text: str | None = None,
    interactive: bool = False,
    extra_env: dict[str, str] | None = None,
    stream_output_to_job: bool = True,
) -> int:
    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("PYTHONIOENCODING", "utf-8")
    env.setdefault("PYTHONUNBUFFERED", "1")
    env.setdefault("LANG", "C.UTF-8")
    env.setdefault("LC_ALL", "C.UTF-8")
    env.setdefault("NO_COLOR", "1")
    if extra_env:
        env.update(extra_env)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    state.update_job(job_id, command=command, log_file=str(log_path), interactive=interactive)

    with log_path.open("w", encoding="utf-8", errors="replace") as log_file:
        log_file.write(f"$ {' '.join(command)}\n\n")
        log_file.flush()
        state.append_log(job_id, f"$ {' '.join(command)}")

        process = subprocess.Popen(
            command,
            cwd=str(cwd),
            env=env,
            stdin=subprocess.PIPE if input_text is not None or interactive else None,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        if process.stdin is not None:
            state.attach_process(job_id, process)

        try:
            writer = None
            if input_text is not None:
                writer = threading.Thread(target=_write_stdin, args=(process, input_text), daemon=True)
                writer.start()

            assert process.stdout is not None
            for raw_line in process.stdout:
                line = _repair_mojibake(raw_line.rstrip("\n"))
                log_file.write(line + "\n")
                log_file.flush()
                if stream_output_to_job:
                    state.append_log(job_id, line)

            if writer:
                writer.join(timeout=5)
            exit_code = process.wait()
            state.append_log(job_id, f"Exit code: {exit_code}")
            log_file.write(f"\nExit code: {exit_code}\n")
            return exit_code
        finally:
            state.detach_process(job_id)


def _repair_mojibake(text: str) -> str:
    markers = ("浣", "鐨", "鏄", "涓", "鍦", "瀹", "闇", "銆", "锛", "鈥")
    if not text or not any(marker in text for marker in markers):
        return text
    repaired = text.encode("gbk", errors="replace").decode("utf-8", errors="replace")
    return repaired if _looks_more_readable(repaired, text) else text


def _looks_more_readable(candidate: str, original: str) -> bool:
    common_words = ("你", "现在", "本地", "工作台", "调用", "项目", "目录", "需求", "测试", "修改", "确认")
    candidate_hits = sum(word in candidate for word in common_words)
    original_hits = sum(word in original for word in common_words)
    if candidate_hits > original_hits:
        return True
    cjk = re.compile(r"[一-鿿]")
    candidate_score = len(cjk.findall(candidate)) - candidate.count("�")
    original_score = len(cjk.findall(original)) - original.count("�")
    return candidate_score >= original_score


def _write_stdin(process: subprocess.Popen, input_text: str) -> None:
    if process.stdin is None:
        return
    try:
        process.stdin.write(input_text)
        process.stdin.flush()
    finally:
        process.stdin.close()


def prepare_input_materials(
    state: WorkbenchState,
    job_id: str,
    requirement: str,
    attachments: list[dict],
    project_root: Path,
    target_url: str = "",
) -> tuple[str, list[dict]]:
    """Save uploaded materials and build text that orchestrator can consume."""
    materials_dir = state.output_dir / "_workbench_uploads" / job_id
    materials_dir.mkdir(parents=True, exist_ok=True)
    materials = []

    for index, attachment in enumerate(attachments, start=1):
        original_name = str(attachment.get("name") or f"attachment_{index}").strip()
        safe_name = _safe_filename(original_name, index)
        extension = Path(safe_name).suffix.lower()
        if extension not in SUPPORTED_ATTACHMENT_EXTENSIONS:
            raise ValueError(f"Unsupported attachment type: {original_name}")

        data_url = str(attachment.get("data") or "")
        encoded = data_url.split(",", 1)[1] if "," in data_url else data_url
        try:
            content = base64.b64decode(encoded, validate=True)
        except Exception as exc:  # noqa: BLE001 - user-facing validation.
            raise ValueError(f"Invalid attachment data: {original_name}") from exc
        if len(content) > 25 * 1024 * 1024:
            raise ValueError(f"Attachment is too large: {original_name}")

        target_path = materials_dir / safe_name
        target_path.write_bytes(content)
        material = {
            "name": original_name,
            "saved_name": safe_name,
            "path": str(target_path),
            "extension": extension,
            "size": len(content),
            "kind": _material_kind(extension),
            "summary": _extract_material_summary(target_path, extension),
        }
        materials.append(material)

    combined_requirement = _build_combined_requirement(requirement, materials, project_root, target_url)
    return combined_requirement, materials


def _safe_filename(name: str, index: int) -> str:
    path_name = Path(name).name or f"attachment_{index}"
    path_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", path_name)
    stem = Path(path_name).stem or f"attachment_{index}"
    suffix = Path(path_name).suffix.lower()
    stem = stem[:80].strip(" ._") or f"attachment_{index}"
    return f"{index:02d}_{stem}{suffix}"


def _material_kind(extension: str) -> str:
    if extension in IMAGE_EXTENSIONS:
        return "image"
    if extension in SPREADSHEET_EXTENSIONS:
        return "spreadsheet"
    return "text"


def _extract_material_summary(path: Path, extension: str) -> str:
    if extension in IMAGE_EXTENSIONS:
        return "图片材料已保存；生成计划时仅记录路径，Codex 执行时会通过 --image 携带给模型。"
    if extension == ".xlsx":
        return _extract_xlsx_text(path)
    if extension == ".xls":
        return "已保存 .xls 文件；当前标准库工作台不能直接解析二进制 xls，后续交给 Codex/人工按文件路径查看。"
    if extension == ".csv":
        return _extract_csv_text(path)
    return _truncate(path.read_text(encoding="utf-8", errors="replace"), 12000)


def _extract_csv_text(path: Path) -> str:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    output = io.StringIO()
    reader = csv.reader(io.StringIO(text))
    for row_index, row in enumerate(reader, start=1):
        if row_index > 80:
            output.write("... CSV 内容已截断，仅展示前 80 行。\n")
            break
        output.write(" | ".join(cell.strip() for cell in row[:30]) + "\n")
    return _truncate(output.getvalue(), 12000)


def _extract_xlsx_text(path: Path) -> str:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        lines = []
        sheet_names = wb.sheetnames
        for sheet_index, sheet_name in enumerate(sheet_names[:8], start=1):
            lines.append(f"[Sheet {sheet_index}: {sheet_name}]")
            sheet = wb[sheet_name]
            for row_index, row in enumerate(sheet.iter_rows(max_row=80, values_only=True), start=1):
                cells = [str(cell).strip() if cell is not None else "" for cell in row[:30]]
                if any(cells):
                    lines.append(" | ".join(cells))
            lines.append("")
        if len(sheet_names) > 8:
            lines.append("... xlsx sheet 数量较多，仅展示前 8 个。")
        wb.close()
        return _truncate("\n".join(lines).strip(), 16000)
    except Exception as exc:  # noqa: BLE001 - keep pipeline usable with a note.
        return f"xlsx 内容抽取失败，文件仍已保存，可由 Codex/人工查看。错误：{exc}"


def _truncate(text: str, limit: int) -> str:
    return text if len(text) <= limit else text[:limit] + "\n... 内容已截断。"


def _build_combined_requirement(
    requirement: str,
    materials: list[dict],
    project_root: Path,
    target_url: str = "",
) -> str:
    parts = ["【项目根目录】\n" + str(project_root)]
    if target_url.strip():
        parts.append("【目标页面 URL】\n" + target_url.strip())
    if requirement.strip():
        parts.append("【需求描述】\n" + requirement.strip())
    if materials:
        lines = ["【需求材料】"]
        for material in materials:
            lines.append(
                f"- {material['name']} ({material['kind']}, {material['size']} bytes): {material['path']}"
            )
        parts.append("\n".join(lines))
    return "\n\n".join(parts).strip()


def write_input_materials(run_dir: Path, materials: list[dict], combined_requirement: str) -> None:
    attachments_dir = run_dir / "attachments"
    md_dir = run_dir / "md"
    json_dir = run_dir / "json"
    copied = []
    for material in materials:
        source = Path(material["path"])
        target = attachments_dir / material["saved_name"]
        target.parent.mkdir(parents=True, exist_ok=True)
        if source.exists():
            shutil.copyfile(source, target)
        item = dict(material)
        item["path"] = str(target)
        item["relative_path"] = f"attachments/{material['saved_name']}"
        copied.append(item)

    payload = {
        "combined_requirement": combined_requirement,
        "materials": copied,
    }
    json_dir.mkdir(parents=True, exist_ok=True)
    (json_dir / "input_materials.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    md_dir.mkdir(parents=True, exist_ok=True)
    (md_dir / "input_materials.md").write_text(
        _build_input_materials_markdown(copied, combined_requirement),
        encoding="utf-8",
    )


def _build_input_materials_markdown(materials: list[dict], combined_requirement: str) -> str:
    lines = ["# 输入材料", "", "## 合并后的需求输入", "", "```text", combined_requirement, "```", ""]
    lines.append("## 附件")
    if not materials:
        lines.append("")
        lines.append("无附件。")
        return "\n".join(lines)
    for material in materials:
        lines.extend([
            "",
            f"### {material['name']}",
            "",
            f"- 类型: `{material['kind']}`",
            f"- 路径: `{material['relative_path']}`",
            f"- 大小: `{material['size']}` bytes",
        ])
    return "\n".join(lines)


def load_input_materials(run_dir: Path) -> list[dict]:
    path = run_dir / "json" / "input_materials.json"
    if not path.exists():
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload.get("materials", [])


def resolve_project_root(state: WorkbenchState, value) -> Path:
    raw_value = str(value or "").strip()
    candidate = Path(raw_value) if raw_value else state.project_root
    try:
        candidate = candidate.expanduser().resolve()
    except OSError as exc:
        raise ValueError(f"Invalid project root: {candidate}") from exc
    if not candidate.exists() or not candidate.is_dir():
        raise ValueError(f"Project root does not exist or is not a directory: {candidate}")
    return normalize_codex_workspace_root(candidate)


def normalize_codex_workspace_root(candidate: Path) -> Path:
    """Use the nearest ancestor with AGENTS.md as the Codex workspace root."""
    if (candidate / "AGENTS.md").exists():
        return candidate
    for parent in candidate.parents:
        if (parent / "AGENTS.md").exists():
            return parent
    return candidate


def normalize_test_path(test_project_root: Path, value: str) -> str:
    raw_value = value.strip()
    if not raw_value:
        raise ValueError("test_path is required")
    path = Path(raw_value)
    if path.is_absolute():
        resolved = path.resolve()
        root = test_project_root.resolve()
        if resolved != root and root not in resolved.parents:
            raise ValueError(f"测试路径必须位于 auto-test 项目内: {resolved}")
        return resolved.relative_to(root).as_posix()

    normalized = raw_value.replace("\\", "/").strip("/")
    parts = PurePosixPath(normalized).parts
    if not normalized or any(part in ("", ".", "..") for part in parts):
        raise ValueError(f"无效测试路径: {raw_value}")
    return normalized


def resolve_test_python(test_project_root: Path, explicit_python: str, fallback_python: str) -> str:
    raw_explicit = str(explicit_python or "").strip()
    if raw_explicit:
        return str(Path(raw_explicit).expanduser().resolve())

    for candidate in (
        test_project_root / "venv" / "Scripts" / "python.exe",
        test_project_root / ".venv" / "Scripts" / "python.exe",
    ):
        if candidate.exists():
            return str(candidate.resolve())

    return fallback_python


def allure_report_url_for_env(env: str) -> str:
    if env == "all":
        return "/allure/allure_report_test/index.html"
    return "/allure/allure_report/index.html"


def _allure_report_item(
    *,
    report_id: str,
    title: str,
    kind: str,
    env: str,
    path: Path,
    url: str,
    deletable: bool = False,
    test_path: str = "",
    test_name: str = "",
    sequence: str = "",
) -> dict:
    return {
        "id": report_id,
        "title": title,
        "kind": kind,
        "env": env,
        "test_path": test_path,
        "test_name": test_name,
        "sequence": sequence,
        "created_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
        "url": url,
        "deletable": deletable,
    }


def _resolve_snapshot_test_path(test_root: Path, file_token: str) -> str:
    if not file_token:
        return ""
    try:
        for path in (test_root / "project").rglob("*.py"):
            rel = path.relative_to(test_root).as_posix()
            if rel.replace("/", "_") == file_token:
                return rel
    except OSError:
        return ""
    return file_token.replace("_", "/")


def _parse_snapshot_dir(test_root: Path, snapshot_dir: Path) -> dict:
    name = snapshot_dir.name
    match = re.match(r"^(?P<stamp>\d+)_(?P<seq>\d+)_(?P<body>.+)$", name)
    if not match:
        return {
            "title": name[:80],
            "test_path": "",
            "test_name": "",
            "sequence": "",
        }

    body = match.group("body")
    parts = body.split("__")
    file_token = parts[0] if parts else ""
    class_name = parts[1] if len(parts) > 1 else ""
    method_name = parts[2] if len(parts) > 2 else ""
    test_name = ".".join(part for part in (class_name, method_name) if part)
    test_path = _resolve_snapshot_test_path(test_root, file_token)
    title = test_name or test_path or name[:80]
    return {
        "title": title,
        "test_path": test_path,
        "test_name": test_name,
        "sequence": match.group("seq"),
    }


def list_allure_reports(state: WorkbenchState) -> list[dict]:
    test_root = resolve_test_project_root(state.project_root)
    report_root = (test_root / "testreport").resolve()
    reports: list[dict] = []

    current_reports = [
        ("current:default", "当前总报告", "current", "test", report_root / "allure_report", "/allure/allure_report/index.html"),
        ("current:test", "当前测试站报告", "current", "test", report_root / "allure_report_test", "/allure/allure_report_test/index.html"),
        (
            "current:production",
            "当前正式站报告",
            "current",
            "production",
            report_root / "allure_report_production",
            "/allure/allure_report_production/index.html",
        ),
    ]
    for report_id, title, kind, env, path, url in current_reports:
        if (path / "index.html").is_file():
            reports.append(
                _allure_report_item(
                    report_id=report_id,
                    title=title,
                    kind=kind,
                    env=env,
                    path=path,
                    url=url,
                )
            )

    snapshot_root = report_root / "report_snapshots"
    if snapshot_root.exists():
        for env_dir in snapshot_root.iterdir():
            if not env_dir.is_dir():
                continue
            for snapshot_dir in env_dir.iterdir():
                if not snapshot_dir.is_dir() or not (snapshot_dir / "index.html").is_file():
                    continue
                rel = snapshot_dir.relative_to(snapshot_root).as_posix()
                parsed = _parse_snapshot_dir(test_root, snapshot_dir)
                reports.append(
                    _allure_report_item(
                        report_id=f"snapshot:{rel}",
                        title=parsed["title"],
                        kind="snapshot",
                        env=env_dir.name,
                        path=snapshot_dir,
                        url="/allure/report_snapshots/" + quote(rel) + "/index.html",
                        deletable=True,
                        test_path=parsed["test_path"],
                        test_name=parsed["test_name"],
                        sequence=parsed["sequence"],
                    )
                )

    return sorted(reports, key=lambda item: item.get("created_at", ""), reverse=True)


def delete_allure_snapshot(state: WorkbenchState, report_id: str) -> dict:
    raw_id = str(report_id or "").strip()
    if not raw_id.startswith("snapshot:"):
        raise ValueError("Only snapshot reports can be deleted")

    test_root = resolve_test_project_root(state.project_root)
    snapshot_root = (test_root / "testreport" / "report_snapshots").resolve()
    rel = PurePosixPath(unquote(raw_id.removeprefix("snapshot:")))
    if len(rel.parts) != 2 or any(part in ("", ".", "..") for part in rel.parts):
        raise ValueError("Invalid snapshot id")

    target = (snapshot_root / Path(*rel.parts)).resolve()
    if snapshot_root not in target.parents or target == snapshot_root:
        raise ValueError("Snapshot path is outside report_snapshots")
    if target.parent.parent != snapshot_root:
        raise ValueError("Only a single snapshot directory can be deleted")
    if not target.is_dir() or not (target / "index.html").is_file():
        raise ValueError("Snapshot report does not exist")

    shutil.rmtree(target)
    return {"ok": True, "deleted": raw_id}


def resolve_output_dir(state: WorkbenchState, value) -> Path:
    raw_value = str(value or "").strip()
    candidate = Path(raw_value) if raw_value else state.output_dir
    try:
        candidate = candidate.expanduser().resolve()
    except OSError as exc:
        raise ValueError(f"Invalid output directory: {candidate}") from exc
    if candidate.exists() and not candidate.is_dir():
        raise ValueError(f"Output path exists but is not a directory: {candidate}")
    return candidate


def apply_settings(state: WorkbenchState, payload: dict) -> dict:
    project_root = resolve_project_root(state, payload.get("project_root"))
    output_dir = resolve_output_dir(state, payload.get("output_dir"))
    model = resolve_model(payload.get("model"), state.model)
    state.update_settings(project_root, output_dir, model=model)
    return {
        "project_root": str(state.project_root),
        "output_dir": str(state.output_dir),
        "model": state.model,
    }


def resolve_model(raw_value: object, default: str = "deepseek-v4-flash") -> str:
    value = str(raw_value or default).strip()
    if value not in DEEPSEEK_MODELS:
        raise ValueError(f"model must be one of {sorted(DEEPSEEK_MODELS)}")
    return value


def select_directory(initial_dir: str = "", title: str = "选择目录") -> str:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception as exc:  # noqa: BLE001 - surface local desktop limitation.
        raise RuntimeError("Folder picker is unavailable because tkinter cannot be loaded.") from exc

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected = filedialog.askdirectory(
            initialdir=initial_dir or None,
            title=title or "选择目录",
            mustexist=True,
        )
    finally:
        root.destroy()
    return selected or ""


def select_file(initial_dir: str = "", title: str = "选择文件") -> str:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception as exc:  # noqa: BLE001 - surface local desktop limitation.
        raise RuntimeError("File picker is unavailable because tkinter cannot be loaded.") from exc

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        selected = filedialog.askopenfilename(
            initialdir=initial_dir or None,
            title=title or "选择文件",
            filetypes=(("Python tests", "*.py"), ("All files", "*.*")),
        )
    finally:
        root.destroy()
    return selected or ""


def start_generation_job(state: WorkbenchState, payload: dict) -> dict:
    requirement = str(payload.get("requirement", "")).strip()
    attachments = payload.get("attachments") or []
    target_url = str(payload.get("target_url", "")).strip()
    if not requirement and not attachments and not target_url:
        raise ValueError("requirement, target_url, or attachment is required")
    apply_settings(state, payload)
    project_root = resolve_project_root(state, payload.get("project_root"))

    review_policy = str(payload.get("review_policy", "auto-review"))
    if review_policy not in REVIEW_POLICIES:
        raise ValueError(f"review_policy must be one of {sorted(REVIEW_POLICIES)}")
    model = resolve_model(payload.get("model"), state.model)

    job = state.create_job("generate")
    job_id = job["id"]
    combined_requirement, materials = prepare_input_materials(
        state,
        job_id,
        requirement,
        attachments,
        project_root,
        target_url,
    )
    thread = threading.Thread(
        target=_run_generation_job,
        args=(
            state,
            job_id,
            combined_requirement,
            materials,
            review_policy,
            model,
            bool(payload.get("full_artifacts")),
        ),
        daemon=True,
    )
    thread.start()
    return job


def _run_generation_job(
    state: WorkbenchState,
    job_id: str,
    requirement: str,
    materials: list[dict],
    review_policy: str,
    model: str,
    full_artifacts: bool,
) -> None:
    before = _existing_run_dirs(state.output_dir)
    temp_log = state.output_dir / "_workbench_logs" / f"pipeline_{job_id}.log"
    requirement_file = state.output_dir / "_workbench_uploads" / job_id / "combined_requirement.txt"
    requirement_file.parent.mkdir(parents=True, exist_ok=True)
    requirement_file.write_text(requirement, encoding="utf-8")
    command = [
        state.python_executable,
        str(state.orchestrator_path),
        "--file",
        str(requirement_file),
        "--output-dir",
        str(state.output_dir),
        "--review-policy",
        review_policy,
    ]
    if full_artifacts:
        command.append("--full-artifacts")

    try:
        exit_code = run_command(
            state,
            job_id,
            command,
            state.orchestrator_path.parent,
            temp_log,
            interactive=True,
            extra_env={"ANTHROPIC_MODEL": model},
        )
        run_dir = _find_new_run_dir(state.output_dir, before)
        if run_dir:
            write_input_materials(run_dir, materials, requirement)
            review = state._read_run_review(run_dir)
            logs_dir = run_dir / "logs"
            logs_dir.mkdir(parents=True, exist_ok=True)
            final_log = logs_dir / "pipeline.log"
            shutil.copyfile(temp_log, final_log)
            state.update_job(
                job_id,
                run_dir=str(run_dir),
                run_url=f"/runs/{quote(run_dir.name)}/index.html",
                review_url=f"/runs/{quote(run_dir.name)}/index.html#doc-md-review-notes-md",
                review_decision=review.get("decision"),
                review_summary=review.get("summary"),
                review_counts=review.get("counts", {}),
                log_file=str(final_log),
            )
        status = "success" if exit_code == 0 and run_dir else "failed"
        error = None if status == "success" else "Pipeline did not finish successfully or no run folder was found."
        state.update_job(job_id, status=status, exit_code=exit_code, error=error)
    except Exception as exc:  # noqa: BLE001 - surface full local failure to the workbench.
        state.append_log(job_id, f"ERROR: {exc}")
        state.update_job(job_id, status="failed", error=str(exc))


def start_codex_job(state: WorkbenchState, payload: dict) -> dict:
    run_ref = str(payload.get("run_dir", "")).strip()
    if not run_ref:
        raise ValueError("run_dir is required")
    apply_settings(state, payload)

    approval_policy = str(payload.get("approval_policy", "on-request"))
    if approval_policy not in APPROVAL_POLICIES:
        raise ValueError(f"approval_policy must be one of {sorted(APPROVAL_POLICIES)}")
    execution_mode = str(payload.get("execution_mode", "analysis"))
    if execution_mode not in EXECUTION_MODES:
        raise ValueError(f"execution_mode must be one of {sorted(EXECUTION_MODES)}")

    return start_codex_execution_job(
        state=state,
        run_dir=state.resolve_run_dir(run_ref),
        project_root=resolve_project_root(state, payload.get("project_root")),
        approval_policy=approval_policy,
        extra_instruction=str(payload.get("extra_instruction", "")).strip(),
        continue_mode=bool(payload.get("continue_mode")),
        execution_mode=execution_mode,
        load_input_materials_func=load_input_materials,
        run_command_func=run_command,
        repair_mojibake_func=_repair_mojibake,
        prepare_authorized_cdp_environment_func=prepare_authorized_cdp_environment,
    )


def start_test_job(state: WorkbenchState, payload: dict) -> dict:
    test_path = str(payload.get("test_path", "")).strip()
    if not test_path:
        raise ValueError("test_path is required")
    env = str(payload.get("env", "test"))
    if env not in ("test", "production", "all"):
        raise ValueError("env must be test, production, or all")
    apply_settings(state, payload)
    project_root = resolve_project_root(state, payload.get("project_root"))
    run_ref = str(payload.get("run_dir", "")).strip()
    run_dir = state.resolve_run_dir(run_ref) if run_ref else None
    evidence_mode = str(payload.get("evidence_mode", "url")).strip() or "url"
    if evidence_mode not in EVIDENCE_MODES:
        raise ValueError(f"evidence_mode must be one of {sorted(EVIDENCE_MODES)}")
    evidence_target_url = str(payload.get("evidence_target_url", "")).strip()
    evidence_selector_filter = str(payload.get("evidence_selector_filter", "")).strip() or None
    evidence_cdp_port = int(payload.get("evidence_cdp_port") or (
        DEFAULT_STORE_DEBUG_PORT if evidence_mode == "store_cdp" else DEFAULT_ELECTRON_DEBUG_PORT
    ))
    job = state.create_job("test")
    job_id = job["id"]
    thread = threading.Thread(
        target=_run_test_job,
        args=(
            state,
            job_id,
            project_root,
            test_path,
            env,
            run_dir,
            evidence_mode,
            evidence_target_url,
            evidence_selector_filter,
            evidence_cdp_port,
        ),
        daemon=True,
    )
    thread.start()
    return job


def _run_test_job(
    state: WorkbenchState,
    job_id: str,
    project_root: Path,
    test_path: str,
    env: str,
    run_dir: Path | None = None,
    evidence_mode: str = "url",
    evidence_target_url: str = "",
    evidence_selector_filter: str | None = None,
    evidence_cdp_port: int = DEFAULT_ELECTRON_DEBUG_PORT,
) -> None:
    log_path = state.output_dir / "_workbench_logs" / f"test_{job_id}.log"
    log_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        test_project_root = resolve_test_project_root(project_root)
        test_python = resolve_test_python(
            test_project_root,
            state.explicit_test_python,
            state.python_executable,
        )
        state.test_python_executable = test_python
        normalized_test_path = normalize_test_path(test_project_root, test_path)
        runner_py = test_project_root / "runner.py"
        command = [test_python, str(runner_py), "--env", env, normalized_test_path]
        report_url = allure_report_url_for_env(env)
        state.update_job(job_id, command=command, log_file=str(log_path), test_env=env)
        state.append_log(job_id, f"$ {' '.join(command)}")
        state.append_log(job_id, f"工作区目录: {project_root}")
        state.append_log(job_id, f"测试项目目录: {test_project_root}")
        state.append_log(job_id, f"平台 Python: {state.python_executable}")
        state.append_log(job_id, f"测试 Python: {test_python}")
        state.append_log(job_id, f"测试路径: {normalized_test_path}")
        state.append_log(job_id, f"环境: {env}")
        state.append_log(job_id, f"完整日志: {log_path}")
        if env == "all":
            state.append_log(job_id, "ALL 模式会生成 /allure/allure_report_test/ 和 /allure/allure_report_production/ 两份报告。")
        exit_code = run_command(
            state,
            job_id,
            command,
            test_project_root,
            log_path,
            stream_output_to_job=True,
        )
        status = "success" if exit_code == 0 else "failed"
        state.update_job(
            job_id,
            status=status,
            exit_code=exit_code,
            allure_report_url=report_url,
        )
        passed = exit_code == 0
        state.append_log(job_id, "")
        state.append_log(job_id, "=" * 40)
        if env == "all" and passed:
            state.append_log(job_id, "ALL 模式执行完成，请分别查看 test / production 两份 Allure 报告。")
        else:
            state.append_log(job_id, f"测试{'全部通过' if passed else '存在失败，请查看 Allure 报告'}")
        state.append_log(job_id, f"Exit code: {exit_code}")
        if not passed:
            maybe_write_failure_evidence_diff(
                state=state,
                job_id=job_id,
                run_dir=run_dir,
                test_path=normalized_test_path,
                log_path=log_path,
                target_url=evidence_target_url,
                mode=evidence_mode,
                selector_filter=evidence_selector_filter,
                cdp_port=evidence_cdp_port,
                browser_channel=config.browser if "config" in globals() else "chrome",
            )
    except Exception as exc:  # noqa: BLE001 - show exact local failure.
        state.append_log(job_id, f"ERROR: {exc}")
        state.update_job(job_id, status="failed", error=str(exc))


def _existing_run_dirs(output_dir: Path) -> set[Path]:
    if not output_dir.exists():
        return set()
    return {path.resolve() for path in output_dir.iterdir() if path.is_dir() and (path / "index.html").exists()}


def _find_new_run_dir(output_dir: Path, before: set[Path]) -> Path | None:
    candidates = []
    for path in output_dir.iterdir():
        if not path.is_dir() or not (path / "index.html").exists():
            continue
        resolved = path.resolve()
        if resolved not in before:
            candidates.append(path)
    if not candidates:
        candidates = [path for path in output_dir.iterdir() if path.is_dir() and (path / "index.html").exists()]
    return max(candidates, key=lambda p: p.stat().st_mtime) if candidates else None


# ---------------------------------------------------------------------------
# FastAPI application
# ---------------------------------------------------------------------------

app = FastAPI(title="自动化测试平台")


def _get_state(request: Request) -> WorkbenchState:
    return request.app.state.workbench


# ---------------------------------------------------------------------------
# API routes
# ---------------------------------------------------------------------------

@app.get("/api/runs")
async def api_list_runs(request: Request):
    return {"runs": _get_state(request).list_runs()}


@app.get("/api/jobs/{job_id}")
async def api_get_job(job_id: str, request: Request):
    job = _get_state(request).get_job(job_id)
    if not job:
        return JSONResponse({"error": "job not found"}, status_code=404)
    return {"job": job}


@app.post("/api/generate")
async def api_generate(request: Request):
    payload = await request.json()
    job = start_generation_job(_get_state(request), payload)
    return {"job": job}


@app.post("/api/codex")
async def api_codex(request: Request):
    payload = await request.json()
    job = start_codex_job(_get_state(request), payload)
    return {"job": job}


@app.post("/api/run-tests")
async def api_run_tests(request: Request):
    payload = await request.json()
    job = start_test_job(_get_state(request), payload)
    return {"job": job}


@app.post("/api/evidence")
async def api_evidence(request: Request):
    payload = await request.json()
    job = start_evidence_capture_job(_get_state(request), payload, apply_settings)
    return {"job": job}


@app.post("/api/settings")
async def api_settings(request: Request):
    payload = await request.json()
    settings = apply_settings(_get_state(request), payload)
    return settings


@app.post("/api/select-directory")
async def api_select_directory(request: Request):
    payload = await request.json()
    selected = select_directory(
        str(payload.get("initial_dir", "")).strip(),
        str(payload.get("title", "选择目录")).strip(),
    )
    return {"path": selected}


@app.post("/api/select-file")
async def api_select_file(request: Request):
    payload = await request.json()
    selected = select_file(
        str(payload.get("initial_dir", "")).strip(),
        str(payload.get("title", "选择文件")).strip(),
    )
    return {"path": selected}


@app.post("/api/runs/delete")
async def api_delete_run(request: Request):
    payload = await request.json()
    run_ref = str(payload.get("run_dir") or payload.get("path") or payload.get("name") or "").strip()
    if not run_ref:
        raise ValueError("run_dir is required")
    _get_state(request).delete_run(run_ref)
    return {"ok": True}


@app.get("/api/allure-reports")
async def api_allure_reports(request: Request):
    return {"reports": list_allure_reports(_get_state(request))}


@app.post("/api/allure-reports/delete")
async def api_delete_allure_report(request: Request):
    payload = await request.json()
    return delete_allure_snapshot(_get_state(request), str(payload.get("id", "")))


@app.post("/api/jobs/{job_id}/input")
async def api_job_input(job_id: str, request: Request):
    payload = await request.json()
    text = str(payload.get("input", ""))
    if not text.strip():
        raise ValueError("input is required")
    job = _get_state(request).send_input(job_id, text)
    return {"job": job}


# ---------------------------------------------------------------------------
# Static file routes
# ---------------------------------------------------------------------------

async def _serve_run_file(state: WorkbenchState, run_name: str, file_path: str) -> Response:
    run_dir = state.resolve_run_dir(unquote(run_name))
    target = (run_dir / unquote(file_path)).resolve()
    if run_dir.resolve() not in target.parents and target != run_dir.resolve():
        return Response(status_code=403)
    if not target.exists() or not target.is_file():
        return Response(status_code=404)
    return FileResponse(target)


@app.get("/")
async def index(request: Request):
    return HTMLResponse(build_index_html(_get_state(request)))


@app.get("/reports")
async def reports_index(request: Request):
    return HTMLResponse(build_index_html(_get_state(request)))


@app.get("/runs/{run_name}")
async def serve_run_root(run_name: str, request: Request):
    return await _serve_run_file(_get_state(request), run_name, "index.html")


@app.get("/runs/{run_name}/{file_path:path}")
async def serve_run_path(run_name: str, file_path: str, request: Request):
    return await _serve_run_file(_get_state(request), run_name, file_path)


@app.get("/allure/{file_path:path}")
async def serve_allure_path(file_path: str, request: Request):
    state = _get_state(request)
    try:
        test_root = resolve_test_project_root(state.project_root)
    except ValueError:
        return Response(status_code=404)
    report_root = (test_root / "testreport").resolve()
    target = (report_root / unquote(file_path)).resolve()
    if report_root not in target.parents and target != report_root:
        return Response(status_code=403)
    if not target.exists() or not target.is_file():
        return Response(status_code=404)
    return FileResponse(target)


# ---------------------------------------------------------------------------
# Exception handlers
# ---------------------------------------------------------------------------

@app.exception_handler(ValueError)
async def _value_error_handler(request: Request, exc: ValueError):
    return JSONResponse({"error": str(exc)}, status_code=400)


@app.exception_handler(FileNotFoundError)
async def _file_not_found_handler(request: Request, exc: FileNotFoundError):
    return JSONResponse({"error": str(exc)}, status_code=404)


@app.exception_handler(Exception)
async def _general_error_handler(request: Request, exc: Exception):
    return JSONResponse({"error": str(exc)}, status_code=500)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _asset_dir() -> Path:
    return Path(__file__).resolve().parent.parent / "assets"


def _read_asset_text(filename: str) -> str:
    return (_asset_dir() / filename).read_text(encoding="utf-8")


def build_index_html(state: WorkbenchState) -> str:
    """Render the workbench shell from assets/workbench.html."""
    content = _read_asset_text("workbench.html")
    asset_version = str(
        max(
            (_asset_dir() / "workbench.css").stat().st_mtime_ns,
            (_asset_dir() / "workbench.js").stat().st_mtime_ns,
        )
    )
    replacements = {
        "{{PORT}}": html.escape(str(state.port)),
        "{{PROJECT_ROOT}}": html.escape(str(state.project_root)),
        "{{OUTPUT_DIR}}": html.escape(str(state.output_dir)),
        "{{ORCHESTRATOR_PATH}}": html.escape(str(state.orchestrator_path)),
        "{{ASSET_VERSION}}": html.escape(asset_version),
    }
    for token, value in replacements.items():
        content = content.replace(token, value)
    return content


def parse_args() -> argparse.Namespace:
    default_orchestrator = Path(__file__).resolve().parent / "orchestrator.py"
    default_codex = shutil.which("codex.cmd") or shutil.which("codex") or "codex.cmd"
    parser = argparse.ArgumentParser(
        description="自动化测试平台本地工作台",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--host", default="127.0.0.1", help="Local bind host")
    parser.add_argument("--port", type=int, default=8765, help="Local bind port")
    parser.add_argument("--project-root", default=str(Path.cwd()), help="Project root passed to codex exec")
    parser.add_argument("--output-dir", default=str(Path.cwd() / "output"), help="Pipeline output directory")
    parser.add_argument("--orchestrator", default=str(default_orchestrator), help="Path to orchestrator.py")
    parser.add_argument("--python", default=sys.executable, help="Python executable used for orchestrator.py")
    parser.add_argument("--test-python", default="", help="Python executable used for auto-test runner.py")
    parser.add_argument("--codex-command", default=default_codex, help="codex executable, usually codex.cmd on Windows")
    parser.add_argument("--open-browser", action="store_true", help="Open the local workbench URL in the default browser")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    project_root = Path(args.project_root)
    test_python_executable = resolve_test_python(project_root, args.test_python, args.python)
    state = WorkbenchState(
        host=args.host,
        port=args.port,
        orchestrator_path=Path(args.orchestrator),
        output_dir=Path(args.output_dir),
        project_root=project_root,
        python_executable=args.python,
        test_python_executable=test_python_executable,
        explicit_test_python=args.test_python,
        codex_command=args.codex_command,
        model=os.environ.get("ANTHROPIC_MODEL", "deepseek-v4-flash"),
    )
    app.state.workbench = state

    assets_dir = _asset_dir()
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=str(assets_dir)), name="assets")

    url = f"http://{state.host}:{state.port}/"
    print(f"自动化测试平台: {url}")
    print(f"Project root: {state.project_root}")
    print(f"Output dir:   {state.output_dir}")
    print("Press Ctrl+C to stop.")
    if args.open_browser:
        webbrowser.open(url)
    uvicorn.run(app, host=state.host, port=state.port, log_level="warning")


if __name__ == "__main__":
    main()
